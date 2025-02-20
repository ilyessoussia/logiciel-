import tkinter as tk
from tkinter import messagebox, filedialog, ttk, simpledialog, Text
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime
import json

class InventoryApp:
    def __init__(self, root):
        self.root = root
        self.root.withdraw()  
        self.current_user = None 
        self.root.title("Gestion des Stocks")
        self.root.geometry("1400x900")  # Increased height to accommodate new fields
        self.root.configure(bg="#f0f0f0")

    # Hide the main window until login is successful
       
        self.create_users_file_if_not_exists()
        self.login_window()

        
        # Data storage
        self.ingredients = []
        self.file_path = "inventory_data.xlsx"  # Default file path
        self.editing_id = None  # Track the ID of the ingredient being edited
        self.history_log = []  # To store changes
        self.history_file = "history_log.txt"  # File to store history


        # Create the file if it doesn't exist
        self.create_file_if_not_exists()

        # Configure styles
        self.configure_styles()

        # GUI Components
        self.create_widgets()
   
        # Create the history file if it doesn’t exist
        self.create_history_file_if_not_exists()

        # Save data when the window is closed
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

        # Set focus to the name entry field when the app starts
        self.name_entry.focus_set()

        self.recipe_history_file = "recipe_history.txt"
        self.create_recipe_history_file_if_not_exists()
        self.create_sales_history_file_if_not_exists()

    def login_window(self):
        """Display the login window."""
        login_win = tk.Toplevel()
        login_win.title("Connexion")
        login_win.geometry("600x400")

        ttk.Label(login_win, text="Nom d'utilisateur:").pack(pady=5)
        username_entry = ttk.Entry(login_win)
        username_entry.pack(pady=5)

        ttk.Label(login_win, text="Mot de passe:").pack(pady=5)
        password_entry = ttk.Entry(login_win, show="*")
        password_entry.pack(pady=5)

        def attempt_login():
            username = username_entry.get().strip()
            password = password_entry.get().strip()

            if self.authenticate_user(username, password):
                self.current_user = username
                self.log_change(f"Connexion réussie par {username}", None, None)
                messagebox.showinfo("Succès", f"Bienvenue {username}!")
                login_win.destroy()
                self.start_main_application()  # ✅ Load data after login
            else:
                messagebox.showerror("Erreur", "Nom d'utilisateur ou mot de passe incorrect!")

        ttk.Button(login_win, text="Connexion", command=attempt_login).pack(pady=10)

        def open_register():
            login_win.destroy()
            self.register_window()

        ttk.Button(login_win, text="Créer un compte", command=open_register).pack(pady=5)

        login_win.protocol("WM_DELETE_WINDOW", self.exit_program)

    def exit_program(self):
        """Exit the application if no login occurs."""
        self.root.quit()  # Ensure proper shutdown

    def start_main_application(self):
        """Initialize the main application after a successful login."""
        self.root.deiconify()  # Show the main window

        # ✅ Now load data
        self.create_file_if_not_exists()
        self.load_excel_data()
        self.load_history()

        # ✅ Show stock alerts after login
        self.check_low_stock()


    def register_window(self):
        """Display the registration window."""
        reg_win = tk.Toplevel()
        reg_win.title("Créer un Compte")
        reg_win.geometry("300x250")

        ttk.Label(reg_win, text="Nom d'utilisateur:").pack(pady=5)
        username_entry = ttk.Entry(reg_win)
        username_entry.pack(pady=5)

        ttk.Label(reg_win, text="Mot de passe:").pack(pady=5)
        password_entry = ttk.Entry(reg_win, show="*")
        password_entry.pack(pady=5)

        ttk.Label(reg_win, text="Confirmer Mot de Passe:").pack(pady=5)
        confirm_entry = ttk.Entry(reg_win, show="*")
        confirm_entry.pack(pady=5)

        def register():
            username = username_entry.get().strip()
            password = password_entry.get().strip()
            confirm_password = confirm_entry.get().strip()

            if password != confirm_password:
                messagebox.showerror("Erreur", "Les mots de passe ne correspondent pas!")
                return

            if self.add_user(username, password):
                messagebox.showinfo("Succès", "Compte créé avec succès!")
                reg_win.destroy()
                self.login_window()
            else:
                messagebox.showerror("Erreur", "Nom d'utilisateur déjà pris!")

        ttk.Button(reg_win, text="S'inscrire", command=register).pack(pady=10)

    def add_user(self, username, password):
        """Add a new user to users.json."""
        with open("users.json", "r", encoding="utf-8") as f:
            users = json.load(f)

        if username in users:
            return False  # Username already exists

        users[username] = password  # Store plain-text for now (consider hashing)

        with open("users.json", "w", encoding="utf-8") as f:
            json.dump(users, f)

        return True

    def authenticate_user(self, username, password):
        """Check if username and password match stored credentials."""
        with open("users.json", "r", encoding="utf-8") as f:
            users = json.load(f)

        return users.get(username) == password


    def configure_styles(self):
        """Configure custom styles for the application."""
        style = ttk.Style()
        style.theme_use("clam")  # Use the 'clam' theme for modern widgets

        # Configure colors and fonts
        style.configure("TFrame", background="#f0f0f0")
        style.configure("TLabel", background="#f0f0f0", font=("Helvetica", 12))
        style.configure("TButton", font=("Helvetica", 10), padding=5)  # Regular button style
        style.configure("Treeview", font=("Helvetica", 11), rowheight=25)
        style.configure("Treeview.Heading", font=("Helvetica", 12, "bold"))
        style.map("Treeview", background=[("selected", "#0078d7")])

        # Add horizontal separators between rows
        style.configure("Treeview", rowheight=25, background="#ffffff", fieldbackground="#ffffff")
        style.configure("Treeview.Separator", background="#cccccc")  # Color of the separator
        style.layout("Treeview.Item", [
            ("Treeitem.padding", {"sticky": "nswe", "children": [
                ("Treeitem.indicator", {"side": "left", "sticky": ""}),
                ("Treeitem.image", {"side": "left", "sticky": ""}),
                ("Treeitem.text", {"side": "left", "sticky": ""}),
            ]}),
            ("Treeitem.separator", {"sticky": "we", "children": [
                ("Treeitem.separator", {"sticky": "we"}),
            ]}),
        ])

    def create_file_if_not_exists(self):
        """Create the Excel file with headers if it doesn't exist."""
        if not os.path.exists(self.file_path):
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Nom", "Quantité", "Unité", "Catégorie", "Seuil Minimum", "Lot", "Date", "Notes"])  # Add headers
            workbook.save(self.file_path)
            messagebox.showinfo("Info", f"Fichier créé : {self.file_path}")
    
    def create_history_file_if_not_exists(self):
        """Create the history log file if it doesn't exist."""
        if not os.path.exists(self.history_file):
            with open(self.history_file, "w", encoding="utf-8") as f:
                 f.write("Historique des modifications:\n")  # Add a header for clarity

    def create_recipe_history_file_if_not_exists(self):
        """Create the recipe history log file if it doesn't exist."""
        if not os.path.exists(self.recipe_history_file):
            with open(self.recipe_history_file, "w", encoding="utf-8") as f:
                f.write("Recipe history:\n") # Add a header for clarity

    def create_sales_history_file_if_not_exists(self):
        """Create the sales history file if it doesn't exist."""
        if not os.path.exists("sales_history.txt"):
            with open("sales_history.txt", "w", encoding="utf-8") as f:
                f.write("Historique des ventes:\n")  # Add a header for clarity

    def create_users_file_if_not_exists(self):
        """Create users.json if missing."""
        if not os.path.exists("users.json"):
            with open("users.json", "w", encoding="utf-8") as f:
                json.dump({}, f)  # Store users as a dictionary



    def create_widgets(self):
        """Create and configure GUI elements."""

        # Dashboard Frame
        self.dashboard_frame = ttk.Frame(self.root)
        self.dashboard_frame.grid(row=0, column=0, columnspan=2, padx=10, pady=10, sticky="ew")

        # Dashboard Cards
        self.total_items_card = self.create_dashboard_card("Articles Totaux", "0")
        self.low_stock_card = self.create_dashboard_card("Articles en Rupture", "0")
        self.categories_card = self.create_dashboard_card("Catégories", "0")

        # Search and Filter Frame
        self.search_filter_frame = ttk.Frame(self.root)
        self.search_filter_frame.grid(row=1, column=0, columnspan=2, padx=10, pady=10, sticky="ew")

        self.search_label = ttk.Label(self.search_filter_frame, text="Rechercher:")
        self.search_label.pack(side="left", padx=5)
        self.search_entry = ttk.Entry(self.search_filter_frame)
        self.search_entry.pack(side="left", fill="x", expand=True, padx=5)
        self.search_entry.bind("<KeyRelease>", self.search_ingredients)  # Search as you type

        self.filter_label = ttk.Label(self.search_filter_frame, text="Filtrer par Catégorie:")
        self.filter_label.pack(side="left", padx=5)
        self.filter_var = tk.StringVar()
        self.filter_dropdown = ttk.Combobox(self.search_filter_frame, textvariable=self.filter_var, values=[])
        self.filter_dropdown.pack(side="left", padx=5)
        self.filter_dropdown.bind("<KeyRelease>", self.filter_ingredients)  # Corrected binding

        self.lot_filter_label = ttk.Label(self.search_filter_frame, text="Filtrer par Lot:")
        self.lot_filter_label.pack(side="left", padx=5)
        self.lot_filter_var = tk.StringVar()
        self.lot_filter_dropdown = ttk.Combobox(self.search_filter_frame, textvariable=self.lot_filter_var, values=[])
        self.lot_filter_dropdown.pack(side="left", padx=5)
        self.lot_filter_dropdown.bind("<KeyRelease>", self.filter_ingredients)

        # File Upload Button
        self.upload_button = ttk.Button(self.root, text="Charger un Fichier Excel", command=self.load_excel_data)
        self.upload_button.grid(row=2, column=0, columnspan=2, padx=10, pady=10, sticky="ew")

        # Add Ingredient Form
        self.form_frame = ttk.LabelFrame(self.root, text="Ajouter/Modifier un Ingredient", padding=10)
        self.form_frame.grid(row=3, column=0, padx=10, pady=10, sticky="nsew")

    # Row 0: Name, Quantity, Unit
        ttk.Label(self.form_frame, text="Nom:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.name_entry = ttk.Entry(self.form_frame)
        self.name_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        ttk.Label(self.form_frame, text="Quantité:").grid(row=0, column=2, padx=5, pady=5, sticky="w")
        self.quantity_entry = ttk.Entry(self.form_frame)
        self.quantity_entry.grid(row=0, column=3, padx=5, pady=5, sticky="ew")

        ttk.Label(self.form_frame, text="Unité:").grid(row=0, column=4, padx=5, pady=5, sticky="w")
        self.unit_entry = ttk.Entry(self.form_frame)
        self.unit_entry.grid(row=0, column=5, padx=5, pady=5, sticky="ew")
  
    # Row 1: Category, Threshold, Lot (All on row 1)
        ttk.Label(self.form_frame, text="Catégorie:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
        self.category_entry = ttk.Entry(self.form_frame)
        self.category_entry.grid(row=1, column=1, padx=5, pady=5, sticky="ew")

        ttk.Label(self.form_frame, text="Seuil Minimum:").grid(row=1, column=2, padx=5, pady=5, sticky="w")
        self.threshold_entry = ttk.Entry(self.form_frame)
        self.threshold_entry.grid(row=1, column=3, padx=5, pady=5, sticky="ew")

        self.lot_label = ttk.Label(self.form_frame, text="Lot:")
        self.lot_label.grid(row=1, column=4, padx=5, pady=5, sticky="w")  # Column 4
        self.lot_entry = ttk.Entry(self.form_frame)
        self.lot_entry.grid(row=1, column=5, padx=5, pady=5, sticky="ew")  # Column 5

    # Row 2: Date, Notes
        self.date_label = ttk.Label(self.form_frame, text="Date:")
        self.date_label.grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.date_entry = ttk.Entry(self.form_frame)  # Consider using a DateEntry widget for better UX
        self.date_entry.grid(row=2, column=1, padx=5, pady=5, sticky="ew")
        self.date_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))  # Default to current date

    # Row 3: Notes
        self.notes_label = ttk.Label(self.form_frame, text="Notes:")
        self.notes_label.grid(row=3, column=0, padx=5, pady=5, sticky="w")
        self.notes_text = Text(self.form_frame, height=3, width=30)
        self.notes_text.grid(row=3, column=1, padx=5, pady=5, sticky="ew")

    # Button Frame (Row 4)
        self.button_frame = ttk.Frame(self.form_frame)
        self.button_frame.grid(row=4, column=0, columnspan=6, pady=10, sticky="ew")  # Span all columns
        self.add_button = ttk.Button(self.button_frame, text="Ajouter Ingredient", command=self.add_or_update_ingredient)
        self.add_button.pack(side="left", padx=5, fill="x", expand=True)
        self.save_button = ttk.Button(self.button_frame, text="Enregistrer", command=self.add_or_update_ingredient)
        self.save_button.pack(side="left", padx=5, fill="x", expand=True)
        self.save_button.pack_forget()  # Hide the "Enregistrer" button initially

    # Configure grid weights for responsiveness in the button frame
        self.button_frame.grid_columnconfigure(0, weight=1)
        self.button_frame.grid_columnconfigure(1, weight=1)

    # Configure column weights for responsiveness in the form frame
        for i in range(6):  # Assuming 6 columns are used in the form
           self.form_frame.grid_columnconfigure(i, weight=1)


        # Inventory List
        self.list_frame = ttk.LabelFrame(self.root, text="Liste des Ingredients", padding=10)
        self.list_frame.grid(row=4, column=0, columnspan=2, padx=10, pady=10, sticky="nsew")

        self.columns = ("Nom", "Quantité", "Unité", "Catégorie", "Seuil Minimum", "Lot", "Date", "Notes")  # Added columns
        self.tree = ttk.Treeview(self.list_frame, columns=self.columns, show="headings")
        # Ensure the frame expands
        self.root.grid_rowconfigure(4, weight=1)
        self.root.grid_columnconfigure(0, weight=1)
        self.root.grid_columnconfigure(1, weight=1)

        for col in self.columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=100, anchor="center", stretch=True)  # Make columns auto-adjust

        self.tree.pack(fill="both", expand=True)

        

        # Action Buttons
        self.action_frame = ttk.Frame(self.root)
        self.action_frame.grid(row=5, column=0, columnspan=2, padx=10, pady=10, sticky="ew")

        self.edit_button = ttk.Button(self.action_frame, text="Modifier", command=self.edit_ingredient)
        self.edit_button.pack(side="left", padx=5, fill="x", expand=True)

        self.delete_button = ttk.Button(self.action_frame, text="Supprimer", command=self.delete_ingredient)
        self.delete_button.pack(side="left", padx=5, fill="x", expand=True)

        self.save_excel_button = ttk.Button(self.action_frame, text="Enregistrer dans Excel", command=self.save_to_excel)
        self.save_excel_button.pack(side="left", padx=5, fill="x", expand=True)

        # Consume Ingredient Button
        self.consume_button = ttk.Button(self.action_frame, text="Consommer Ingredient", command=self.consume_ingredient)
        self.consume_button.pack(side="left", padx=5, fill="x", expand=True)

        # article sortie
        self.sortie_button = ttk.Button(self.action_frame, text="Sortie Article", command=self.sortie_article)
        self.sortie_button.pack(side="left", padx=5, fill="x", expand=True)


        self.create_recipe_button = ttk.Button(self.action_frame, text="Créer une Recette", command=self.create_recipe)
        self.create_recipe_button.pack(side="left", padx=5, fill="x", expand=True)

        # History Log Button (Optional, for viewing the log)
        self.history_button = ttk.Button(self.action_frame, text="Voir l'Historique", command=self.view_history)
        self.history_button.pack(side="left", padx=5, fill="x", expand=True)

        # Configure grid weights for responsiveness
        self.root.grid_columnconfigure(0, weight=1)
        self.root.grid_columnconfigure(1, weight=1)
        self.root.grid_rowconfigure(4, weight=1)  # Allow the inventory list to expand

    def create_dashboard_card(self, title, value):
        """Create a dashboard card with a title and value."""
        card = ttk.Frame(self.dashboard_frame, style="TFrame")
        card.pack(side="left", fill="both", expand=True, padx=10, pady=10)

        title_label = ttk.Label(card, text=title, style="TLabel")
        title_label.pack(pady=5)

        value_label = ttk.Label(card, text=value, style="TLabel", font=("Helvetica", 16, "bold"))
        value_label.pack(pady=5)

        return value_label

    def update_dashboard(self):
        """Update dashboard values."""
        total_items = len(self.ingredients)
        low_stock_items = sum(1 for i in self.ingredients if float(i["quantity"]) < float(i["min_threshold"]))
        categories = len(set(i["category"] for i in self.ingredients))

        self.total_items_card.config(text=str(total_items))
        self.low_stock_card.config(text=str(low_stock_items))
        self.categories_card.config(text=str(categories))

        # Highlight the low stock card if there are low stock items
        if low_stock_items > 0:
            self.low_stock_card.config(background="#ffcccc", foreground="red")  # Light red background, red text
        else:
            self.low_stock_card.config(background="#f0f0f0", foreground="black")  # Reset to default

    def load_excel_data(self):
        """Load data from the Excel file."""
        try:
            workbook = load_workbook(self.file_path)
            sheet = workbook.active
            self.ingredients = []
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
                if row[0] is not None:  # Check if the first cell in the row is not empty
                    self.ingredients.append({
                        "id": len(self.ingredients) + 1,  # Assign a unique ID
                        "name": row[0],
                        "quantity": row[1],
                        "unit": row[2],
                        "category": row[3],
                        "min_threshold": row[4],
                        "lot": row[5] if len(row) > 5 else "",  # Handle missing columns
                        "date": row[6] if len(row) > 6 else datetime.now().strftime("%Y-%m-%d"),
                        "notes": row[7] if len(row) > 7 else ""
                    })
            self.update_treeview()
            self.update_dashboard()
            self.update_filter_dropdown()
            messagebox.showinfo("Succès", "Données chargées avec succès!")

        except FileNotFoundError:
            messagebox.showerror("Erreur", "Fichier Excel non trouvé.")
        except Exception as e:
            messagebox.showerror("Erreur", f"Échec du chargement du fichier : {e}")

    def update_treeview(self):
        """Update the treeview with the current ingredient data."""
        self.tree.delete(*self.tree.get_children())  # Clear the current treeview
        for ingredient in self.ingredients:
            if float(ingredient["quantity"]) < float(ingredient["min_threshold"]):
                self.tree.insert("", "end", values=(
                    ingredient["name"],
                    ingredient["quantity"],
                    ingredient["unit"],
                    ingredient["category"],
                    ingredient["min_threshold"],
                    ingredient["lot"],
                    ingredient["date"],
                    ingredient["notes"]
                ), tags=("low_stock", "bold"))  # Add a tag for styling, and 'bold' tag
            else:
                self.tree.insert("", "end", values=(
                    ingredient["name"],
                    ingredient["quantity"],
                    ingredient["unit"],
                    ingredient["category"],
                    ingredient["min_threshold"],
                    ingredient["lot"],
                    ingredient["date"],
                    ingredient["notes"]
                ), tags=("bold",))  # Add the 'bold' tag

        # Configure the tag for low stock items
        self.tree.tag_configure("low_stock", background="#ffcccc")  # Light red background
        self.tree.tag_configure("bold", font=("Helvetica", 11, "bold"))  # Configure 'bold' tag

    def search_ingredients(self, event=None):
        """Filter ingredients by search term."""
        search_term = self.search_entry.get().strip().lower()
        filtered_ingredients = [i for i in self.ingredients if search_term in i["name"].lower()]
        self.update_treeview_with_data(filtered_ingredients)

    def filter_ingredients(self, event=None):
        """Filter ingredients by category and lot number."""
        category = self.filter_var.get().strip().lower()
        lot = self.lot_filter_var.get().strip()
      
        if not category and not lot:
            self.update_treeview()  # Reload all ingredients
            return  # Exit the function early 
 
        filtered_ingredients = self.ingredients

    # Filter by category if selected
        if category:
            filtered_ingredients = [i for i in filtered_ingredients if i["category"].lower() == category]

    # Filter by lot if selected
        if lot:
            filtered_ingredients = [i for i in filtered_ingredients if i["lot"] == lot]

        self.update_treeview_with_data(filtered_ingredients)

    def update_treeview_with_data(self, data):
        """Update the treeview with provided data."""
        self.tree.delete(*self.tree.get_children())
        for ingredient in data:
            if float(ingredient["quantity"]) < float(ingredient["min_threshold"]):
                self.tree.insert("", "end", values=(
                    ingredient["name"],
                    ingredient["quantity"],
                    ingredient["unit"],
                    ingredient["category"],
                    ingredient["min_threshold"],
                    ingredient["lot"],
                    ingredient["date"],
                    ingredient["notes"]
                ), tags=("low_stock", "bold"))  # Add a tag for styling, and 'bold' tag
            else:
                self.tree.insert("", "end", values=(
                    ingredient["name"],
                    ingredient["quantity"],
                    ingredient["unit"],
                    ingredient["category"],
                    ingredient["min_threshold"],
                    ingredient["lot"],
                    ingredient["date"],
                    ingredient["notes"]
                ), tags=("bold",))  # Add the 'bold' tag

    def update_filter_dropdown(self):
        """Update the filter dropdown with available categories."""
        categories = list(set(i["category"] for i in self.ingredients))
        lots = list(set(i["lot"] for i in self.ingredients))
        self.filter_dropdown["values"] = categories
        self.lot_filter_dropdown["values"] = lots

    def check_low_stock(self):
        """Check for low stock items and display a notification."""
        low_stock_items = [i for i in self.ingredients if float(i["quantity"]) < float(i["min_threshold"])]
        if low_stock_items:
            messagebox.showwarning("Alerte de Stock", f"{len(low_stock_items)} articles sont en rupture de stock!")

    def validate_inputs(self):
        """Validate inputs before adding or editing an ingredient."""
        name = self.name_entry.get().strip()
        quantity = self.quantity_entry.get().strip()
        unit = self.unit_entry.get().strip()
        category = self.category_entry.get().strip()
        threshold = self.threshold_entry.get().strip()
        lot = self.lot_entry.get().strip()  # Get lot number
        date_str = self.date_entry.get().strip()

        if not all([name, quantity, unit, category, threshold, lot, date_str]):  # Check lot and date
            messagebox.showwarning("Erreur de Saisie", "Tous les champs sont requis!")
            return False

        try:
            quantity = float(quantity)
            threshold = float(threshold)
            datetime.strptime(date_str, "%Y-%m-%d")  # Validate date format
        except ValueError:
            messagebox.showwarning("Erreur de Saisie", "Quantité et Seuil Minimum doivent être des nombres, et la date doit être au format AAAA-MM-JJ!")
            return False

        if quantity < 0 or threshold < 0:
            messagebox.showwarning("Erreur de Saisie", "Quantité et Seuil Minimum doivent être positifs!")
            return False

        if self.editing_id is None and any(i["name"].lower() == name.lower() and i["lot"] == lot for i in self.ingredients):
            messagebox.showwarning("Erreur de Saisie", "Un ingrédient avec ce nom existe déjà!")
            return False

        return True

    def add_or_update_ingredient(self):
        """Add or update an ingredient based on whether editing_id is set."""
        if not self.validate_inputs():
            return

        name = self.name_entry.get().strip()
        quantity = float(self.quantity_entry.get().strip())
        unit = self.unit_entry.get().strip()
        category = self.category_entry.get().strip()
        threshold = float(self.threshold_entry.get().strip()) if self.threshold_entry.get().strip() else 0
        lot = self.lot_entry.get().strip()
        date_str = self.date_entry.get().strip()
        notes = self.notes_text.get("1.0", tk.END).strip()  # Get notes from text widget

        if self.editing_id is not None:
            # Update existing ingredient
            for ingredient in self.ingredients:
                if ingredient["id"] == self.editing_id:
                    old_values = ingredient.copy()  # Log old values
                    ingredient.update({
                        "name": name,
                        "quantity": quantity,
                        "unit": unit,
                        "category": category,
                        "min_threshold": threshold,
                        "lot": lot,
                        "date": date_str,
                        "notes": notes
                    })
                    self.log_change(f"Ingredient modifié: {name}", old_values, ingredient)
                    break
        else:
            # Add new ingredient
            new_ingredient = {
                "id": len(self.ingredients) + 1,  # Assign a unique ID
                "name": name,
                "quantity": quantity,
                "unit": unit,
                "category": category,
                "min_threshold": threshold,
                "lot": lot,
                "date": date_str,
                "notes": notes
            }
            self.ingredients.append(new_ingredient)
            self.log_change(f"Nouvel ingrédient ajouté: {name}", None, new_ingredient)  # Log addition
            self.record_purchase(new_ingredient)  # Record the purchase

        self.update_treeview()
        self.update_dashboard()
        self.update_filter_dropdown()
        self.clear_form()
        self.show_add_button()  # Return to add button state
        self.editing_id = None
        self.save_to_excel()

    def edit_ingredient(self):
        """Load the selected ingredient's data into the form for editing."""
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("Erreur de Sélection", "Veuillez sélectionner un ingrédient à modifier!")
            return

        item = self.tree.item(selected_item)
        values = item["values"]

        # Find the ingredient by name and populate the form
        for ingredient in self.ingredients:
            if ingredient["name"] == values[0]:
                self.editing_id = ingredient["id"]  # Set the editing ID
                self.name_entry.delete(0, tk.END)
                self.name_entry.insert(0, ingredient["name"])
                self.quantity_entry.delete(0, tk.END)
                self.quantity_entry.insert(0, ingredient["quantity"])
                self.unit_entry.delete(0, tk.END)
                self.unit_entry.insert(0, ingredient["unit"])
                self.category_entry.delete(0, tk.END)
                self.category_entry.insert(0, ingredient["category"])
                self.threshold_entry.delete(0, tk.END)
                self.threshold_entry.insert(0, ingredient["min_threshold"])
                self.lot_entry.delete(0, tk.END)
                self.lot_entry.insert(0, ingredient["lot"])
                self.date_entry.delete(0, tk.END)
                self.date_entry.insert(0, ingredient["date"])
                self.notes_text.delete("1.0", tk.END)  # Clear existing text
                self.notes_text.insert("1.0", ingredient["notes"])

                self.show_save_button()  # Show "Save" button instead of "Add"
                break

    def delete_ingredient(self):
        """Delete the selected ingredient."""
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("Erreur de Sélection", "Veuillez sélectionner un ingrédient à supprimer!")
            return

        item = self.tree.item(selected_item)
        values = item["values"]
        ingredient_name = values[0]

        if messagebox.askyesno("Confirmation", f"Êtes-vous sûr de vouloir supprimer {ingredient_name} ?"):
            # Find the ingredient by name and remove it
            for ingredient in self.ingredients:
                if ingredient["name"] == ingredient_name:
                    self.ingredients.remove(ingredient)
                    self.log_change(f"Ingrédient supprimé: {ingredient_name}", ingredient, None)  # Log deletion
                    break

            self.update_treeview()
            self.update_dashboard()
            self.update_filter_dropdown()
            self.save_to_excel()

    def consume_ingredient(self):
        """Consume an ingredient by reducing its quantity."""
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("Erreur de Sélection", "Veuillez sélectionner un ingrédient à consommer!")
            return

        item = self.tree.item(selected_item)
        values = item["values"]
        ingredient_name = values[0]

        # Find the ingredient by name
        for ingredient in self.ingredients:
            if ingredient["name"] == ingredient_name:
                # Ask the user for the quantity to consume
                quantity_to_consume = simpledialog.askfloat("Consommer Ingrédient", f"Quantité à consommer pour {ingredient_name}:",
                                                            initialvalue=0.0)

                if quantity_to_consume is not None and quantity_to_consume > 0:
                    if quantity_to_consume <= ingredient["quantity"]:
                        old_values = ingredient.copy()
                        ingredient["quantity"] -= quantity_to_consume
                        self.log_change(f"Ingredient consommé: {ingredient_name}", old_values, ingredient)  # Log consumption
                        self.update_treeview()
                        self.update_dashboard()
                        self.save_to_excel()
                    else:
                        messagebox.showwarning("Erreur", "Quantité à consommer est supérieure à la quantité disponible.")
                    break
                else:
                    messagebox.showwarning("Erreur", "Veuillez entrer une quantité valide à consommer.")
                break

    def sortie_article(self):
        """Open a window for selling articles and log the transaction."""
        sale_window = tk.Toplevel(self.root)
        sale_window.title("Sortie Article")
        sale_window.geometry("800x500")

        # Buyer Name
        ttk.Label(sale_window, text="Nom de l'Acheteur:").pack(pady=5)
        buyer_entry = ttk.Entry(sale_window)
        buyer_entry.pack(pady=5)

        # Table for selecting articles
        columns = ("Nom", "Quantité Disponible", "Quantité Vendue")
        sale_tree = ttk.Treeview(sale_window, columns=columns, show="headings", selectmode="extended")

        for col in columns:
            sale_tree.heading(col, text=col)
            sale_tree.column(col, width=200, anchor="center")

        sale_tree.pack(fill="both", expand=True, padx=10, pady=10)

        # Populate with 'Article' category items
        article_items = [item for item in self.ingredients if item["category"].lower() == "article"]
        for item in article_items:
            sale_tree.insert("", "end", values=(item["name"], item["quantity"], "0"))

        # Edit quantity on double-click
        def edit_quantity(event):
            selected_item = sale_tree.selection()
            if not selected_item:
                return

            item_id = selected_item[0]
            values = sale_tree.item(item_id, "values")

            quantity_available = float(values[1])
            quantity_sold = simpledialog.askfloat("Quantité à vendre", f"Entrez la quantité à vendre (Max {quantity_available}):",
                                                minvalue=0, maxvalue=quantity_available)

            if quantity_sold is not None:
                sale_tree.item(item_id, values=(values[0], values[1], str(quantity_sold)))

        sale_tree.bind("<Double-1>", edit_quantity)

        # Save transaction
        def save_sale():
            buyer_name = buyer_entry.get().strip()
            if not buyer_name:
                messagebox.showerror("Erreur", "Veuillez entrer un nom d'acheteur.")
                return

            selected_sales = []
            for item in sale_tree.get_children():
                name, available, sold = sale_tree.item(item, "values")
                try:
                    sold = float(sold)
                    if sold > 0:
                        selected_sales.append({"name": name, "quantity": sold})
                except ValueError:
                    continue

            if not selected_sales:
                messagebox.showerror("Erreur", "Veuillez sélectionner au moins un article.")
                return

            # Deduct from inventory
            for sale in selected_sales:
                for item in self.ingredients:
                    if item["name"] == sale["name"]:
                        item["quantity"] -= sale["quantity"]
                        self.log_change(f"Vente: {sale['name']}", {"quantity": item["quantity"] + sale["quantity"]}, item)

            # Log the sale
            sale_record = f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} -  Vendeur: {self.current_user}, Acheteur: {buyer_name}, Articles: {selected_sales}\n"
            with open("sales_history.txt", "a", encoding="utf-8") as f:
                f.write(sale_record)

            messagebox.showinfo("Succès", "Vente enregistrée avec succès!")
            self.update_treeview()
            self.update_dashboard()
            self.save_to_excel()
            sale_window.destroy()

        # Save Button
        save_button = ttk.Button(sale_window, text="Enregistrer la Vente", command=save_sale)
        save_button.pack(pady=10, padx=10, ipadx=10, ipady=5, fill="x", expand=True)




    def create_recipe(self):
        """Handle the creation of a new recipe and deduct ingredients."""

    # Create a new Toplevel window for the recipe creation process
        recipe_window = tk.Toplevel(self.root)
        recipe_window.title("Créer une Nouvelle Recette")
        recipe_window.geometry("800x500")

    # Add a Label and Entry for the Recipe Name
        recipe_name_label = ttk.Label(recipe_window, text="Nom de la Recette:")
        recipe_name_label.pack(pady=5)
        recipe_name_entry = ttk.Entry(recipe_window)
        recipe_name_entry.pack(pady=5)

    # Create a Treeview with an editable column for required quantities
        recipe_columns = ("Nom", "Quantité Disponible", "Quantité Requise", "Lot")
        recipe_tree = ttk.Treeview(recipe_window, columns=recipe_columns, show="headings")
    
        for col in recipe_columns:
            recipe_tree.heading(col, text=col)
            recipe_tree.column(col, width=150, anchor="center")

        recipe_tree.pack(fill="both", expand=True, padx=10, pady=10)

    # Populate the treeview with available ingredients
        for ingredient in self.ingredients:
            recipe_tree.insert("", "end", values=(ingredient["name"], ingredient["quantity"], 0, ingredient["lot"]))

        def edit_quantity(event):
            selected_item = recipe_tree.selection()
            if not selected_item:
                return
        
            item = selected_item[0]
            values = recipe_tree.item(item, "values")
        
            ingredient_name = values[0]
            available_quantity = values[1]
        
        # Ask user for required quantity
            required_quantity = simpledialog.askfloat("Quantité Requise", 
                                                  f"Entrez la quantité requise pour {ingredient_name} (Max {available_quantity}):",
                                                  minvalue=0.0, maxvalue=float(available_quantity))

            if required_quantity is not None:
                recipe_tree.item(item, values=(ingredient_name, available_quantity, str(required_quantity), values[3]))

    # Bind double-click event to allow input
        recipe_tree.bind("<Double-1>", edit_quantity)

    # Function to save the recipe and deduct ingredients
        def save_recipe():
            recipe_name = recipe_name_entry.get().strip()
            if not recipe_name:
                messagebox.showerror("Erreur", "Veuillez entrer un nom de recette.")
                return

            recipe_ingredients = []

        # Iterate over selected ingredients in the tree
            for item in recipe_tree.get_children():
                ingredient_name, available_quantity, quantity_required, lot = recipe_tree.item(item, "values")
            
                try:
                    quantity_required = float(quantity_required)  # Convert to float
                    available_quantity = float(available_quantity)
                except ValueError:
                    messagebox.showerror("Erreur", f"Quantité invalide pour {ingredient_name}. Veuillez entrer un nombre.")
                    return

                if quantity_required > 0:
                    if quantity_required > available_quantity:
                        messagebox.showerror("Erreur", f"Quantité insuffisante de {ingredient_name} en stock.")
                        return
                    recipe_ingredients.append({
                        "name": ingredient_name,
                        "quantity": quantity_required,
                        "lot": lot
                    })

        # Deduct Ingredients from Inventory
            for ingredient_data in recipe_ingredients:
                ingredient_name = ingredient_data["name"]
                quantity_required = ingredient_data["quantity"]

                for ingredient in self.ingredients:
                    if ingredient["name"] == ingredient_name:
                        ingredient["quantity"] -= quantity_required
                        self.log_change(f"Recette: {recipe_name} - Consommé {quantity_required} {ingredient['lot']} de {ingredient_name}",
                                    {"name": ingredient_name, "quantity": ingredient["quantity"] + quantity_required},
                                    ingredient)
                        break

        # Update UI and save changes
            self.update_treeview()
            self.update_dashboard()
            self.save_to_excel()
            messagebox.showinfo("Succès", f"Recette '{recipe_name}' enregistrée et ingrédients déduits avec succès!")

        # Save the recipe in recipe history
            with open(self.recipe_history_file, "a", encoding="utf-8") as f:
                f.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - Créateur: {self.current_user}, Recette: {recipe_name} - Ingredients: {recipe_ingredients}\n")

            recipe_window.destroy()

    # Add a "Save Recipe" Button
        save_button = ttk.Button(recipe_window, text="Enregistrer la Recette", command=save_recipe)
        save_button.pack(pady=10, padx=10, ipadx=10, ipady=5, fill="x", expand=True)


    def save_to_excel(self):
        """Save the current ingredient data to the Excel file."""
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Nom", "Quantité", "Unité", "Catégorie", "Seuil Minimum", "Lot", "Date", "Notes"])  # Headers
            for ingredient in self.ingredients:
                sheet.append([
                    ingredient["name"],
                    ingredient["quantity"],
                    ingredient["unit"],
                    ingredient["category"],
                    ingredient["min_threshold"],
                    ingredient["lot"],
                    ingredient["date"],
                    ingredient["notes"]
                ])
            workbook.save(self.file_path)
            messagebox.showinfo("Succès", "Données enregistrées dans Excel!")
        except Exception as e:
            messagebox.showerror("Erreur", f"Échec de l'enregistrement dans Excel : {e}")

    def record_purchase(self, ingredient):
        """Record the purchase of a new ingredient (you can extend this to a separate purchases table/log)."""
        purchase_details = f"Achat de {ingredient['quantity']} {ingredient['unit']} de {ingredient['name']} (Lot: {ingredient['lot']}) le {ingredient['date']}"
        self.log_change("Achat enregistré", None, {"details": purchase_details})
        messagebox.showinfo("Achat Enregistré", purchase_details)

    def log_change(self, action, old_values, new_values):
        """Log changes to history and save to file."""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_entry = {
            "timestamp": timestamp,
            "user": self.current_user,
            "action": action,
            "old_values": old_values,
            "new_values": new_values
    }
        self.history_log.append(log_entry)

    # Convert log to text and save it
        log_text = f"{timestamp} - {action}\n"
        if old_values:
            log_text += f"  Anciennes valeurs: {old_values}\n"
        if new_values:
            log_text += f"  Nouvelles valeurs: {new_values}\n"
        log_text += "\n"
 
    # Append the log to the history file
        with open(self.history_file, "a", encoding="utf-8") as f:
            f.write(log_text)
    
    def load_history(self):
        """Load history from the text file into the history log list."""
        if os.path.exists(self.history_file):
            with open(self.history_file, "r", encoding="utf-8") as f:
                lines = f.readlines()
        
        # Store history in list format (only for displaying later)
        current_log = []
        for line in lines:
            if line.strip():  # Ignore empty lines
                current_log.append(line.strip())
        self.history_log = current_log  # Store for display in view_history()

    def view_history(self):
        """Display the history log in a new window."""
        history_window = tk.Toplevel(self.root)
        history_window.title("Historique des Modifications")
        history_text = tk.Text(history_window, height=20, width=100)
        history_text.pack(padx=10, pady=10)
        if os.path.exists(self.history_file):
            with open(self.history_file, "r", encoding="utf-8") as f:
               history_text.insert(tk.END, f.read())
        history_text.config(state=tk.DISABLED)  # Make read-only

    def clear_form(self):
        """Clear the input form."""
        self.name_entry.delete(0, tk.END)
        self.quantity_entry.delete(0, tk.END)
        self.unit_entry.delete(0, tk.END)
        self.category_entry.delete(0, tk.END)
        self.threshold_entry.delete(0, tk.END)
        self.lot_entry.delete(0, tk.END)
        self.date_entry.delete(0, tk.END)
        self.date_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))  # Reset to current date
        self.notes_text.delete("1.0", tk.END)

    def show_save_button(self):
        """Show the Save button and hide the Add button."""
        self.add_button.pack_forget()
        self.save_button.pack(side="left", padx=5, fill="x", expand=True)

    def show_add_button(self):
        """Show the Add button and hide the Save button."""
        self.save_button.pack_forget()
        self.add_button.pack(side="left", padx=5, fill="x", expand=True)

    def on_close(self):
        """Save data to Excel and close the application."""
        if messagebox.askokcancel("Quitter", "Voulez-vous enregistrer les modifications et quitter ?"):
            self.save_to_excel()
            self.root.destroy()

if __name__ == "__main__":
    root = tk.Tk()
    app = InventoryApp(root)
    root.mainloop()
